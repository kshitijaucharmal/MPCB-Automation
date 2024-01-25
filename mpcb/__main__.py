import sys
from PyQt6 import QtWidgets, QtGui
from openpyxl import load_workbook, Workbook
import re
import os

class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Company Data Extractor")
        self.setMinimumSize(600, 400)

        self.central_widget = QtWidgets.QWidget()
        self.setCentralWidget(self.central_widget)

        self.layout = QtWidgets.QGridLayout()
        self.central_widget.setLayout(self.layout)

        self.testfile_label = QtWidgets.QLabel("Test File Location:")
        self.testfile_textbox = QtWidgets.QLineEdit()
        self.browse_testfile_button = QtWidgets.QPushButton("Browse")
        self.browse_testfile_button.clicked.connect(self.browse_testfile)

        self.company_name_label = QtWidgets.QLabel("Company Name:")
        self.company_name_textbox = QtWidgets.QLineEdit()

        self.generate_button = QtWidgets.QPushButton("Generate")
        self.generate_button.clicked.connect(self.generate_file)

        self.layout.addWidget(self.testfile_label, 0, 0)
        self.layout.addWidget(self.testfile_textbox, 0, 1)
        self.layout.addWidget(self.browse_testfile_button, 0, 2)

        self.layout.addWidget(self.company_name_label, 1, 0)
        self.layout.addWidget(self.company_name_textbox, 1, 1)
        self.layout.addWidget(self.generate_button, 2, 0, 1, 2)

    def browse_testfile(self):
        filename, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Select Test File", "", "Excel Files (*.xlsx)")

        if filename:
            self.testfile_textbox.setText(filename)

    def generate_file(self):
        testfile_location = self.testfile_textbox.text()
        company_name = self.company_name_textbox.text()

        if not testfile_location:
            QtWidgets.QMessageBox.warning(self, "Error", "Please select a test file.")
            return

        new_wb = Workbook()
        new_sheet = new_wb.active

        workbook = load_workbook(os.path.realpath(testfile_location))

        def write_to_file(row, name):
            data = []
            for cell in row:
                if cell.value:
                    data.append(cell.value)
            data.append(name)
            new_sheet.append(data)

        for sheet in workbook.worksheets:
            for i, cell in enumerate(sheet["B"]):
                val = cell.value
                try:
                    if val and re.search(company_name, val, re.IGNORECASE):
                        write_to_file(sheet[i+1], sheet.title)
                except Exception:
                    continue

        self.new_wb = new_wb
        self.company_name = company_name

        self.save_file()

    def save_file(self):
        filename, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Save File", f"{self.company_name}.xlsx")

        if filename:
            self.new_wb.save(filename=filename)
            QtWidgets.QMessageBox.information(self, "Success", "File saved successfully.")

def main():
    app = QtWidgets.QApplication(sys.argv)
    window = MainWindow()
    window.show()
    app.exec()

if __name__ == "__main__":
    main()
